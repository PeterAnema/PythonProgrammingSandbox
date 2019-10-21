import openpyxl

wb = openpyxl.Workbook()

ws = wb.active

for col in range(1,40):
    column_letter = openpyxl.utils.get_column_letter(col)
    for row in range(1,600):
        ws.cell(col,row).value = '%s%s'%(column_letter,row) + '!!! CBS !!!'
        
ws = wb.create_sheet()

ws.title = 'openpyxl'

ws['F5'] = 'CBS !!!! use it!!'

wb.save(filename='using_openpyxl.xlsx')

#----------------------------------------------------

wb = openpyxl.load_workbook(filename = 'using_openpyxl.xlsx')
print(wb.sheetnames)
ws = wb['openpyxl']
print(ws['F5'].value)
